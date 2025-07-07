import io
import os
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from models.evaluation import EvaluationSubmission, db
from openpyxl import Workbook

evaluation_bp = Blueprint('evaluation', __name__)

# Use environment variable for admin secret (more secure)
ADMIN_SECRET = os.getenv('ADMIN_SECRET', 'kirokhela')


def verify_admin(secret):
    """Simple verification function"""
    return secret == ADMIN_SECRET


@evaluation_bp.route('/submit', methods=['POST'])
def submit_evaluation():
    """Handle form submission"""
    try:
        # Handle both form data and JSON
        if request.content_type and 'application/json' in request.content_type:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['participant_name', 'main_team']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False, 
                    'error': f'Missing required field: {field}'
                }), 400

        # Helper function to safely convert to int
        def safe_int(value):
            try:
                return int(value) if value and str(value).strip() else None
            except (ValueError, TypeError):
                return None

        # Create new evaluation submission
        submission = EvaluationSubmission(
            participant_name=data.get('participant_name', '').strip(),
            main_team=data.get('main_team', '').strip(),
            sub_team=data.get('sub_team', '').strip(),
            program_rating=safe_int(data.get('program_rating')),
            program_pros=data.get('program_pros', '').strip(),
            program_cons=data.get('program_cons', '').strip(),
            leaders_rating=safe_int(data.get('leaders_rating')),
            leaders_pros=data.get('leaders_pros', '').strip(),
            leaders_cons=data.get('leaders_cons', '').strip(),
            games_rating=safe_int(data.get('games_rating')),
            games_pros=data.get('games_pros', '').strip(),
            games_cons=data.get('games_cons', '').strip(),
            goal_delivery_rating=safe_int(data.get('goal_delivery_rating')),
            goal_delivery_pros=data.get('goal_delivery_pros', '').strip(),
            goal_delivery_cons=data.get('goal_delivery_cons', '').strip(),
            logo_rating=safe_int(data.get('logo_rating')),
            logo_pros=data.get('logo_pros', '').strip(),
            logo_cons=data.get('logo_cons', '').strip(),
            gift_rating=safe_int(data.get('gift_rating')),
            gift_pros=data.get('gift_pros', '').strip(),
            gift_cons=data.get('gift_cons', '').strip(),
            secretary_rating=safe_int(data.get('secretary_rating')),
            secretary_pros=data.get('secretary_pros', '').strip(),
            secretary_cons=data.get('secretary_cons', '').strip(),
            media_rating=safe_int(data.get('media_rating')),
            media_pros=data.get('media_pros', '').strip(),
            media_cons=data.get('media_cons', '').strip(),
            emergency_rating=safe_int(data.get('emergency_rating')),
            emergency_pros=data.get('emergency_pros', '').strip(),
            emergency_cons=data.get('emergency_cons', '').strip(),
            kitchen_rating=safe_int(data.get('kitchen_rating')),
            kitchen_pros=data.get('kitchen_pros', '').strip(),
            kitchen_cons=data.get('kitchen_cons', '').strip(),
            finance_rating=safe_int(data.get('finance_rating')),
            finance_pros=data.get('finance_pros', '').strip(),
            finance_cons=data.get('finance_cons', '').strip(),
            custody_rating=safe_int(data.get('custody_rating')),
            custody_pros=data.get('custody_pros', '').strip(),
            custody_cons=data.get('custody_cons', '').strip(),
            purchase_rating=safe_int(data.get('purchase_rating')),
            purchase_pros=data.get('purchase_pros', '').strip(),
            purchase_cons=data.get('purchase_cons', '').strip(),
            transportation_rating=safe_int(data.get('transportation_rating')),
            transportation_pros=data.get('transportation_pros', '').strip(),
            transportation_cons=data.get('transportation_cons', '').strip(),
            general_suggestions=data.get('general_suggestions', '').strip()
        )

        db.session.add(submission)
        db.session.commit()

        return jsonify({
            'success': True, 
            'message': 'تم إرسال التقييم بنجاح',
            'id': submission.id
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"Error in submit_evaluation: {str(e)}")  # Add logging
        return jsonify({
            'success': False, 
            'error': 'حدث خطأ أثناء إرسال التقييم'
        }), 500


@evaluation_bp.route('/download', methods=['GET'])
def download_excel():
    """Download all evaluations as Excel file - requires admin secret"""
    try:
        # Check for admin secret in query parameters
        secret = request.args.get('secret')
        if not verify_admin(secret):
            return jsonify({'error': 'Unauthorized access'}), 401

        # Get all submissions
        submissions = EvaluationSubmission.query.all()

        if not submissions:
            return jsonify({'error': 'No data available'}), 404

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "التقييمات"

        # Define headers
        headers = [
            'ID', 'تاريخ الإرسال', 'الاسم الثلاثي', 'الفريق الأساسي', 'الفريق الفرعي',
            'تقييم البرنامج', 'إيجابيات البرنامج', 'سلبيات البرنامج',
            'تقييم توزيع القادة', 'إيجابيات توزيع القادة', 'سلبيات توزيع القادة',
            'تقييم الألعاب', 'إيجابيات الألعاب', 'سلبيات الألعاب',
            'تقييم توصيل الهدف', 'إيجابيات توصيل الهدف', 'سلبيات توصيل الهدف',
            'تقييم الشعار', 'إيجابيات الشعار', 'سلبيات الشعار',
            'تقييم الهدايا', 'إيجابيات الهدايا', 'سلبيات الهدايا',
            'تقييم السكرتارية', 'إيجابيات السكرتارية', 'سلبيات السكرتارية',
            'تقييم الميديا', 'إيجابيات الميديا', 'سلبيات الميديا',
            'تقييم الإسعافات', 'إيجابيات الإسعافات', 'سلبيات الإسعافات',
            'تقييم المطبخ', 'إيجابيات المطبخ', 'سلبيات المطبخ',
            'تقييم المالية', 'إيجابيات المالية', 'سلبيات المالية',
            'تقييم العهدة', 'إيجابيات العهدة', 'سلبيات العهدة',
            'تقييم المشتريات', 'إيجابيات المشتريات', 'سلبيات المشتريات',
            'تقييم الانتقالات', 'إيجابيات الانتقالات', 'سلبيات الانتقالات',
            'الاقتراحات العامة'
        ]

        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add data rows
        for row, submission in enumerate(submissions, 2):
            data = [
                submission.id,
                submission.submission_date.strftime('%Y-%m-%d %H:%M:%S') if submission.submission_date else '',
                submission.participant_name or '',
                submission.main_team or '',
                submission.sub_team or '',
                submission.program_rating or '',
                submission.program_pros or '',
                submission.program_cons or '',
                submission.leaders_rating or '',
                submission.leaders_pros or '',
                submission.leaders_cons or '',
                submission.games_rating or '',
                submission.games_pros or '',
                submission.games_cons or '',
                submission.goal_delivery_rating or '',
                submission.goal_delivery_pros or '',
                submission.goal_delivery_cons or '',
                submission.logo_rating or '',
                submission.logo_pros or '',
                submission.logo_cons or '',
                submission.gift_rating or '',
                submission.gift_pros or '',
                submission.gift_cons or '',
                submission.secretary_rating or '',
                submission.secretary_pros or '',
                submission.secretary_cons or '',
                submission.media_rating or '',
                submission.media_pros or '',
                submission.media_cons or '',
                submission.emergency_rating or '',
                submission.emergency_pros or '',
                submission.emergency_cons or '',
                submission.kitchen_rating or '',
                submission.kitchen_pros or '',
                submission.kitchen_cons or '',
                submission.finance_rating or '',
                submission.finance_pros or '',
                submission.finance_cons or '',
                submission.custody_rating or '',
                submission.custody_pros or '',
                submission.custody_cons or '',
                submission.purchase_rating or '',
                submission.purchase_pros or '',
                submission.purchase_cons or '',
                submission.transportation_rating or '',
                submission.transportation_pros or '',
                submission.transportation_cons or '',
                submission.general_suggestions or ''
            ]

            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with current date
        filename = f"evaluation_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error in download_excel: {str(e)}")  # Add logging
        return jsonify({'error': str(e)}), 500


@evaluation_bp.route('/stats', methods=['GET'])
def get_stats():
    """Get basic statistics - requires admin secret"""
    try:
        secret = request.args.get('secret')
        if not verify_admin(secret):
            return jsonify({'error': 'Unauthorized access'}), 401

        total_submissions = EvaluationSubmission.query.count()

        return jsonify({
            'total_submissions': total_submissions,
            'message': f'Total submissions: {total_submissions}'
        }), 200

    except Exception as e:
        print(f"Error in get_stats: {str(e)}")  # Add logging
        return jsonify({'error': str(e)}), 500


# Health check endpoint
@evaluation_bp.route('/health', methods=['GET'])
def health_check():
    """Simple health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    }), 200
