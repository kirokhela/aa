import os
from openpyxl import Workbook, load_workbook

# Path to the Excel file
EXCEL_PATH = "evaluation_data.xlsx"

# Fields to save
FIELDS = [
    "participant_name", "main_team", "sub_team",
    "program_rating", "program_pros", "program_cons",
    "leaders_rating", "leaders_pros", "leaders_cons",
    "games_rating", "games_pros", "games_cons",
    "goal_delivery_rating", "goal_delivery_pros", "goal_delivery_cons",
    "logo_rating", "logo_pros", "logo_cons",
    "gift_rating", "gift_pros", "gift_cons",
    "secretary_rating", "secretary_pros", "secretary_cons",
    "media_rating", "media_pros", "media_cons",
    "emergency_rating", "emergency_pros", "emergency_cons",
    "kitchen_rating", "kitchen_pros", "kitchen_cons",
    "finance_rating", "finance_pros", "finance_cons",
    "custody_rating", "custody_pros", "custody_cons",
    "purchase_rating", "purchase_pros", "purchase_cons",
    "transportation_rating", "transportation_pros", "transportation_cons",
]

def save_to_excel(data):
    file_exists = os.path.exists(EXCEL_PATH)
    
    if file_exists:
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(FIELDS)  # Header row

    row = [data.get(field, "") for field in FIELDS]
    sheet.append(row)
    wb.save(EXCEL_PATH)
